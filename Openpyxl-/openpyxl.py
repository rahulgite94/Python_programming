#!/usr/bin/env python
# coding: utf-8

# You get an Excel spreadsheet that has two worksheets, "patient info" and "records"<br>
# 
# Each worksheet has data and visualizations and formulas
# Each row in each worksheet represents information associated with a single person.. ​<br>
# Worksheet "patient info" has a "patient id" column; ​<br>
# Worksheet "records" has a "p_id" column​<br>
# The "patient id" values in "patient info" tab are the same variable as "p_id" in "records". Some people have information on both sheets, while other people have information only on one sheet.
# Use openpyxl to copy information about patients from "records" to "patient info"​<br>. Submit a notebook that reads the Excel spreadsheet and produces a separate spreadsheet with the following modifications:​<br>
# 
# For each patient identified by "p_id" listed in the "records" sheet that don't exist in "patient info," create a new row in "patient info". The row from "records" should be copied to the row in "patient info."
# For each patient listed in the "records" sheet where "p_id" appears "patient info" under "patient id," copy the data in the row from records to "patient info" sheet
# Make no changes to the visualizations that exist in each worksheet​<br>
# Make no changes to the data in the "records"​ sheet<br><br>
# Write your changes to a new .xlsx file (don't overwrite the original)<br>

# #### we will be using openpyxl

# In[1]:


import openpyxl


# #### Load the Workbook

# In[2]:


wb= openpyxl.load_workbook('week_05_homework_XLSX_openpyxl.xlsx')


# In[3]:


print(wb.sheetnames)


# In[4]:


print(wb.active)


# Getting the sheet by name and storing it in a variable

# In[5]:


patientInfo=wb['patient info']


# coping the patientInfo sheet to a new patient info copy

# In[6]:


wb.copy_worksheet(patientInfo)


# Saving the worksheet

# In[7]:


wb.save('week_05_homework_XLSX_openpyxl.xlsx')


# Again loading the workbook

# In[8]:


wb= openpyxl.load_workbook('week_05_homework_XLSX_openpyxl.xlsx')


# In[9]:


print(wb.sheetnames)


# #### getting the whole sheet into variables
# now we will refer patient info Copy worksheet as copyPatient

# In[10]:


records=wb['records']
copyPatient=wb['patient info Copy']


# Getting the patient_id column from patient info Copy

# In[11]:


patientId=[]
first_column = copyPatient['A']
for x in range(len(first_column)):
    patientId.append(first_column[x].value)


# In[12]:


patientId[4]


# Getting the p_id from records

# In[13]:


p_id=[]
first_column_records=records['A']
for x in range(len(first_column_records)):
    p_id.append(first_column_records[x].value)


# Getting all the values which are in p_id but not in patient_id

# In[14]:


notCommonPatient=[]
for x in p_id:
    if x not in patientId:
        notCommonPatient.append(x)


# In[15]:


commonPatient=[]
for x in p_id:
    if x in patientId:
        commonPatient.append(x)


# In[16]:


commonPatient


# Getting the maximum rows for records and patient info

# In[17]:


max_row_records=records.max_row
max_row_copyPatient=copyPatient.max_row
max_row_copyPatient


# Getting the max column for records and patient info

# In[18]:


max_col_records=records.max_column
max_col_copyPatient=copyPatient.max_column
max_col_copyPatient


# Activating hte patient info Copy as we will be coping data to it

# In[19]:


wb.active=2


# In[20]:


wb.active


# #### Coping the common id's which are in records and patient info
# copyRows= rows of copyPatient <br>
# copyCol=column of copyPatient<br>
# recRows = Rows of Records<br>
# recCol = Column of Record<br>

# In[21]:


#copyCol=max_col_copyPatient       
for copyRows in range(1,max_row_copyPatient+1):# iterating copypatient
    if(str(copyPatient.cell(row=copyRows, column=1).value)) in str(commonPatient): # cheking for id in common list
        copyCol=max_col_copyPatient # Assigning the 6 value to copycol
        for recRows in range(1,max_row_records+1):# interating records to find the above copy Patient value
            if(str(copyPatient.cell(row=copyRows, column=1).value)) == (str(records.cell(row=recRows, column=1).value)):#finding the position in records
                for recCol in range(1,max_col_records+1):
                    copyCol=copyCol+1
                    copyPatient.cell(row=copyRows, column=copyCol).value=records.cell(row=recRows,column=recCol).value # Putting all the values of that columns into copypatient


# Initializing variable with max value

# In[22]:


copyRows= max_row_copyPatient
copyCol=max_col_copyPatient


# #### Coping the non common id's value from Records to copyPatient

# In[26]:


for recRows in range(1,max_row_records+1):
    if (str(records.cell(row=recRows, column=1).value)) in str(notCommonPatient):#checking form not common list
        copyRows=copyRows+1
        copyCol=max_col_copyPatient
        for recCol in range(1,max_col_records+1):
            copyCol=copyCol+1
            copyPatient.cell(row=copyRows, column=copyCol).value=records.cell(row=recRows,column=recCol).value 
            # putting the not common values in the row=774 and column = 7 and soo on
            # can also put the values in row=774, column=1, but to symmetric of record data kept it on 7


# In[24]:


copyPatient.max_column


# #### Saving the Workbook

# In[25]:


wb.save('week_05_homework_XLSX_openpyxl.xlsx')

